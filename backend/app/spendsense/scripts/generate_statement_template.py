#!/usr/bin/env python3
"""
Generate the Monytix statement upload Excel template.
Run from repo root: python -m app.spendsense.scripts.generate_statement_template
Output: backend/static/Monytix_Statement_Template.xlsx (and optionally copied to frontend public)
"""

from __future__ import annotations

import asyncio
import sys
from pathlib import Path

# Add backend to path when run as script
backend_path = Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(backend_path))

import pandas as pd  # type: ignore[import-untyped]
import asyncpg  # type: ignore[reportMissingImports]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]

from app.core.config import get_settings

# Column headers that the spendsense parser recognizes (see common.COLUMN_ALIASES)
HEADERS = [
    "Date",           # → txn_date
    "Amount",         # → amount (positive number)
    "Direction",      # → debit or credit
    "Description",   # → description_raw
    "Merchant",       # → merchant_raw (optional; can be inferred from description)
    "Category",       # → category_code (MANDATORY)
    "Subcategory",    # → subcategory_code (MANDATORY)
    "Currency",       # → optional, default INR
    "Reference",      # → raw_txn_id (optional)
]

SAMPLE_ROWS = [
    ("2025-01-15", 450.00, "debit", "UPI-Swiggy-ORDER123", "Swiggy", "dining", "zomato", "INR", "TXN001"),
    ("2025-01-16", 1200.50, "debit", "Amazon Pay-ORDER456", "Amazon", "shopping", "amazon", "INR", "TXN002"),
    ("2025-01-20", 50000.00, "credit", "Salary credit Jan 2025", "", "income", "", "INR", "SAL001"),
]

INSTRUCTIONS = """
Monytix Statement Upload Template
=================================

Required columns:
• Date         – Transaction date (YYYY-MM-DD or DD/MM/YYYY)
• Amount       – Amount as a positive number
• Direction    – "debit" (money out) or "credit" (money in)
• Category     – Select from dropdown (MANDATORY)
• Subcategory  – Select from dropdown (MANDATORY, depends on Category)

Optional columns:
• Description  – Narration/details (used to infer merchant if Merchant is empty)
• Merchant     – Payee or merchant name (e.g. Swiggy, Amazon)
• Currency     – e.g. INR (defaults to INR if empty)
• Reference    – Your reference or transaction ID

Tips:
• Keep the header row as-is so the file is recognized.
• Category and Subcategory are MANDATORY - use the dropdowns to select values.
• Subcategory options change based on the selected Category.
• You can add more rows below the sample data.
• Save as .xlsx or .csv and upload via SpendSense → Upload.
• Dates can be in YYYY-MM-DD or DD/MM/YYYY format.
""".strip()


async def fetch_categories_and_subcategories() -> tuple[list[tuple[str, str]], dict[str, list[tuple[str, str]]]]:
    """Fetch categories and subcategories from database."""
    settings = get_settings()
    conn = await asyncpg.connect(
        str(settings.postgres_dsn),
        statement_cache_size=0,
    )
    try:
        # Fetch categories
        category_rows = await conn.fetch("""
            SELECT category_code, category_name
            FROM spendsense.dim_category
            WHERE active = TRUE
            ORDER BY display_order, category_name
        """)
        categories = [(row['category_code'], row['category_name']) for row in category_rows]

        # Fetch subcategories grouped by category
        subcat_rows = await conn.fetch("""
            SELECT category_code, subcategory_code, subcategory_name
            FROM spendsense.dim_subcategory
            WHERE active = TRUE
            ORDER BY category_code, display_order, subcategory_name
        """)
        subcategories_by_category: dict[str, list[tuple[str, str]]] = {}
        for row in subcat_rows:
            cat_code = row['category_code']
            if cat_code not in subcategories_by_category:
                subcategories_by_category[cat_code] = []
            subcategories_by_category[cat_code].append(
                (row['subcategory_code'], row['subcategory_name'])
            )

        return categories, subcategories_by_category
    finally:
        await conn.close()


def main() -> None:
    out_dir = backend_path / "static"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "Monytix_Statement_Template.xlsx"

    # Fetch categories and subcategories from database
    print("Fetching categories and subcategories from database...")
    categories, subcategories_by_category = asyncio.run(fetch_categories_and_subcategories())
    print(f"Found {len(categories)} categories and {sum(len(subs) for subs in subcategories_by_category.values())} subcategories")

    # Create DataFrame
    df = pd.DataFrame(SAMPLE_ROWS, columns=HEADERS)

    # Create workbook with openpyxl directly to have better control
    from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create helper sheets FIRST (before main data sheet)
    # Category lookup sheet - store as "Category Name (category_code)" for user-friendly display
    cat_sheet = wb.create_sheet("_CategoryLookup")
    cat_sheet.append(["category_display"])  # More user-friendly header
    category_codes = []
    for cat_code, cat_name in categories:
        # Store as "Category Name (category_code)" so users see names but codes are parseable
        display_value = f"{cat_name} ({cat_code})"
        cat_sheet.append([display_value])
        category_codes.append(cat_code)
    
    # Subcategory lookup sheet - store as "Subcategory Name (subcategory_code)"
    all_subcats = []
    for subcats in subcategories_by_category.values():
        all_subcats.extend(subcats)
    # Deduplicate by code while preserving name
    unique_subcats_dict = {}
    for sub_code, sub_name in all_subcats:
        if sub_code not in unique_subcats_dict:
            unique_subcats_dict[sub_code] = sub_name
    
    if unique_subcats_dict:
        subcat_sheet = wb.create_sheet("_SubcategoryLookup")
        subcat_sheet.append(["subcategory_display"])  # More user-friendly header
        unique_subcat_codes = sorted(unique_subcats_dict.keys())
        for sub_code in unique_subcat_codes:
            sub_name = unique_subcats_dict[sub_code]
            display_value = f"{sub_name} ({sub_code})"
            subcat_sheet.append([display_value])
    
    # Now create main sheets
    ws = wb.create_sheet("Transactions")
    
    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write sample data
    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Instructions sheet
    inst_sheet = wb.create_sheet("Instructions")
    for row_idx, line in enumerate(INSTRUCTIONS.split("\n"), start=1):
        inst_sheet.cell(row=row_idx, column=1, value=line)

    # Open on Transactions sheet so users see the main template first (not CategoryLookup)
    wb.active = ws

    # Category dropdown (column F) - use helper sheet reference
    category_formula = f'_CategoryLookup!$A$2:$A${len(category_codes) + 1}'
    category_dv = DataValidation(
        type="list",
        formula1=category_formula,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Category",
        error="Please select a category from the dropdown list.",
    )
    category_dv.add("F2:F1048576")  # Apply to all rows (Excel max rows)
    ws.add_data_validation(category_dv)

    # Subcategory dropdown (column G)
    if unique_subcat_codes:
        subcategory_formula = f'_SubcategoryLookup!$A$2:$A${len(unique_subcat_codes) + 1}'
        subcategory_dv = DataValidation(
            type="list",
            formula1=subcategory_formula,
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Subcategory",
            error="Please select a subcategory from the dropdown list.",
        )
        subcategory_dv.add("G2:G1048576")
        ws.add_data_validation(subcategory_dv)

    # Direction dropdown (column C)
    direction_dv = DataValidation(
        type="list",
        formula1='"debit,credit"',
        allow_blank=False,
    )
    direction_dv.add("C2:C1048576")
    ws.add_data_validation(direction_dv)

    # Keep helper sheets visible (Excel sometimes has issues with hidden sheets in formulas)
    # They're prefixed with "_" so they're less prominent but still accessible
    # Users can hide them manually if desired

    # Save with proper Excel format - ensure we're saving as xlsx
    wb.save(out_path)
    print(f"Generated: {out_path}")

    # Optionally copy to frontend public for direct download
    import shutil
    frontend_public = backend_path.parent / "mony_mvp" / "public"
    if frontend_public.is_dir():
        dest = frontend_public / "Monytix_Statement_Template.xlsx"
        shutil.copy2(out_path, dest)
        print(f"Copied to: {dest}")


if __name__ == "__main__":
    main()
